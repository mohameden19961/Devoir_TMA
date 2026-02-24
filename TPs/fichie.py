from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
import random

# ── Palette ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
ALT_ROW    = "EBF3FB"
WHITE      = "FFFFFF"
GOLD       = "F4B942"
GRAY_BRD   = "B8CCE4"
GREEN_DARK = "375623"
GREEN_FILL = "C6EFCE"
ORANGE_DARK= "9C5700"
ORANGE_FILL= "FFEB9C"
RED_DARK   = "9C0006"
RED_FILL   = "FFC7CE"

def bd(color=GRAY_BRD):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bf(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def hf(size=11, color=WHITE):
    return Font(name="Calibri", size=size, bold=True, color=color)

def fl(c): return PatternFill("solid", fgColor=c)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header_cell(cell, text, bg=MID_BLUE, size=11):
    cell.value     = text
    cell.font      = hf(size, WHITE)
    cell.fill      = fl(bg)
    cell.alignment = ctr()
    cell.border    = bd(MID_BLUE)

def style_label(cell, text, bg=ALT_ROW):
    cell.value     = text
    cell.font      = bf(10, bold=True, color=DARK_BLUE)
    cell.fill      = fl(bg)
    cell.alignment = lft()
    cell.border    = bd()

def style_value(cell, formula, fmt="0.00", color=MID_BLUE):
    cell.value         = formula
    cell.number_format = fmt
    cell.font          = Font(name="Calibri", size=10, bold=True, color=color)
    cell.fill          = fl(WHITE)
    cell.alignment     = ctr()
    cell.border        = bd()

def section_banner(ws, row, col, span, text, bg=DARK_BLUE):
    c1 = get_column_letter(col)
    c2 = get_column_letter(col + span - 1)
    ws.merge_cells(f"{c1}{row}:{c2}{row}")
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    c.fill = fl(bg)
    c.alignment = ctr()
    ws.row_dimensions[row].height = 28

# ── Reference Data ────────────────────────────────────────────────────────────
filieres  = ["RSS", "DWM", "DSI"]
salles    = ["101", "102", "103", "201", "202", "203", "301", "302", "303"]
incidents = ["Coupure Wi-Fi totale", "Connexion instable", "DNS indisponible",
             "Latence élevée", "Déconnexion répétée"]

random.seed(42)
rows = []
for i in range(1, 21):
    f  = random.choice(filieres)
    s  = random.choice(salles)
    inc= random.choice(incidents)
    d  = random.randint(5, 45)
    t  = random.randint(1, 8)
    ni = round(random.uniform(12, 20), 1)
    p  = round(random.uniform(0.5, 5.0), 1)
    nf = round(max(0, ni - p), 1)
    rows.append([f"2400{i:02d}", f, s, inc, d, t, ni, nf])

DATA_FIRST = 3
DATA_LAST  = DATA_FIRST + len(rows) - 1  # 22

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 1 – Données  (colonnes A:I)
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Données"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A3"
ws.sheet_properties.tabColor = DARK_BLUE

# Titre principal
ws.merge_cells("A1:I1")
c = ws["A1"]
c.value     = "📊  ANALYSE DES INCIDENTS DE CONNEXION – SupNum  |  2025-2026"
c.font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
c.fill      = fl(DARK_BLUE)
c.alignment = ctr()
ws.row_dimensions[1].height = 38

# En-têtes colonnes
hdrs = ["ID Étudiant","Filière","Salle","Type d'Incident",
        "Durée (min)","Nb Tentatives","Note Initiale","Note Finale","Impact"]
widths = [14, 10, 9, 26, 13, 15, 14, 13, 10]
for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
    style_header_cell(ws.cell(row=2, column=ci), h)
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[2].height = 30

# Lignes de données
for ri, row_data in enumerate(rows, DATA_FIRST):
    bg = ALT_ROW if ri % 2 == 0 else WHITE
    for ci, val in enumerate(row_data, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = bf(10)
        c.fill = fl(bg)
        c.alignment = lft() if ci == 4 else ctr()
        c.border = bd()
    # Colonne Impact = formule (Q7 – obligatoire)
    imp = ws.cell(row=ri, column=9, value=f"=G{ri}-H{ri}")
    imp.number_format = "0.0"
    imp.font  = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
    imp.fill  = fl(bg)
    imp.alignment = ctr()
    imp.border = bd()

# Échelle de couleur sur Impact
ws.conditional_formatting.add(f"I{DATA_FIRST}:I{DATA_LAST}",
    ColorScaleRule(start_type="num", start_value=0,   start_color="63BE7B",
                   mid_type="num",   mid_value=2.5,   mid_color="FFEB84",
                   end_type="num",   end_value=5,     end_color="F8696B"))

# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 2 – Analyse  (Questions 1 à 6)
# ══════════════════════════════════════════════════════════════════════════════
wa = wb.create_sheet("Analyse")
wa.sheet_view.showGridLines = False
wa.sheet_properties.tabColor = MID_BLUE
wa.column_dimensions["A"].width = 40
wa.column_dimensions["B"].width = 16
wa.column_dimensions["C"].width = 3   # spacer
wa.column_dimensions["D"].width = 26
wa.column_dimensions["E"].width = 16

# Titre
wa.merge_cells("A1:E1")
c = wa["A1"]
c.value     = "📈  PARTIE 2 – ANALYSE DES INCIDENTS  (Questions 1 à 6)"
c.font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
c.fill      = fl(DARK_BLUE)
c.alignment = ctr()
wa.row_dimensions[1].height = 38

# ── Q1 : Durée moyenne ───────────────────────────────────────────────────────
section_banner(wa, 3, 1, 2, "Q1 – Durée moyenne des coupures", MID_BLUE)
style_label(wa["A4"], "Durée moyenne des coupures (minutes)")
style_value(wa["B4"], "=AVERAGE(Données!E3:E22)", "0.00")

# ── Q2 : Min / Max ──────────────────────────────────────────────────────────
section_banner(wa, 6, 1, 2, "Q2 – Durée minimale et maximale", MID_BLUE)
style_label(wa["A7"], "Durée minimale (minutes)", WHITE)
style_value(wa["B7"], "=MIN(Données!E3:E22)", "0")

style_label(wa["A8"], "Durée maximale (minutes)")
style_value(wa["B8"], "=MAX(Données!E3:E22)", "0")

# ── Q3 : Nombre d'incidents par type ────────────────────────────────────────
section_banner(wa, 10, 1, 2, "Q3 – Nombre total d'incidents par type", MID_BLUE)
style_header_cell(wa.cell(row=11, column=1), "Type d'Incident",   LIGHT_BLUE, 10)
wa.cell(row=11, column=1).font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
style_header_cell(wa.cell(row=11, column=2), "Nb Incidents", LIGHT_BLUE, 10)
wa.cell(row=11, column=2).font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)

for j, inc in enumerate(incidents, 12):
    bg = ALT_ROW if j % 2 == 0 else WHITE
    c = wa.cell(row=j, column=1, value=inc)
    c.font = bf(10); c.fill = fl(bg); c.alignment = lft(); c.border = bd()
    v = wa.cell(row=j, column=2, value=f'=COUNTIF(Données!D$3:D$22,A{j})')
    v.number_format = "0"
    v.font = Font(name="Calibri", size=10, bold=True, color=MID_BLUE)
    v.fill = fl(bg); v.alignment = ctr(); v.border = bd()

# ── Q4 : Salle ayant le plus d'incidents ────────────────────────────────────
section_banner(wa, 18, 1, 2, "Q4 – Salle ayant le plus d'incidents", MID_BLUE)
# helper table (hidden helper col A/B rows 20-28)
style_header_cell(wa.cell(row=19, column=1), "Salle", LIGHT_BLUE, 10)
wa.cell(row=19, column=1).font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
style_header_cell(wa.cell(row=19, column=2), "Nb Incidents", LIGHT_BLUE, 10)
wa.cell(row=19, column=2).font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
for j, sal in enumerate(salles, 20):
    bg = ALT_ROW if j % 2 == 0 else WHITE
    c = wa.cell(row=j, column=1, value=sal)
    c.font = bf(10); c.fill = fl(bg); c.alignment = ctr(); c.border = bd()
    v = wa.cell(row=j, column=2, value=f'=COUNTIF(Données!C$3:C$22,A{j})')
    v.number_format = "0"
    v.font = Font(name="Calibri", size=10, bold=True, color=MID_BLUE)
    v.fill = fl(bg); v.alignment = ctr(); v.border = bd()

# Résultat salle max
style_label(wa["A30"], "🏆  Salle la plus touchée", GOLD)
wa["A30"].font = Font(name="Calibri", size=10, bold=True, color=WHITE)
res_sal = wa["B30"]
res_sal.value = f'=INDEX(A20:A{19+len(salles)},MATCH(MAX(B20:B{19+len(salles)}),B20:B{19+len(salles)},0))'
res_sal.font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
res_sal.fill = fl(LIGHT_BLUE); res_sal.alignment = ctr(); res_sal.border = bd()
wa.row_dimensions[30].height = 22

# ── Q5 : Impact moyen sur les notes ─────────────────────────────────────────
section_banner(wa, 32, 1, 2, "Q5 – Impact moyen sur les notes", MID_BLUE)
style_label(wa["A33"], "Impact moyen sur les notes")
style_value(wa["B33"], "=AVERAGE(Données!I3:I22)", "0.00")

# ── Q6 : Étudiants ayant perdu > 2 points ───────────────────────────────────
section_banner(wa, 35, 1, 2, "Q6 – Étudiants ayant perdu plus de 2 points", MID_BLUE)
style_label(wa["A36"], "Nombre d'étudiants avec Impact > 2 points")
style_value(wa["B36"], '=COUNTIF(Données!I3:I22,">2")', "0")

for r in [4,7,8,33,36]:
    wa.row_dimensions[r].height = 22

# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 3 – Gravité & Filtre  (Questions 7 + 8)
# ══════════════════════════════════════════════════════════════════════════════
wg = wb.create_sheet("Gravité & Filtre")
wg.sheet_view.showGridLines = False
wg.freeze_panes = "A3"
wg.sheet_properties.tabColor = "70AD47"

# Titre
wg.merge_cells("A1:J1")
c = wg["A1"]
c.value     = "🚦  Q7 – GRAVITÉ DES INCIDENTS  |  Q8 – FILTRE COMBINÉ (Salle + Type)"
c.font      = Font(name="Calibri", size=13, bold=True, color=WHITE)
c.fill      = fl(DARK_BLUE)
c.alignment = ctr()
wg.row_dimensions[1].height = 38

# En-têtes
hdrs_g = ["ID Étudiant","Filière","Salle","Type d'Incident",
          "Durée (min)","Nb Tentatives","Note Initiale","Note Finale","Impact","Gravité"]
widths_g = [14, 10, 9, 26, 13, 15, 14, 13, 10, 12]
for ci, (h, w) in enumerate(zip(hdrs_g, widths_g), 1):
    style_header_cell(wg.cell(row=2, column=ci), h)
    wg.column_dimensions[get_column_letter(ci)].width = w
wg.row_dimensions[2].height = 30

# Données + formules
for ri, row_data in enumerate(rows, DATA_FIRST):
    bg = ALT_ROW if ri % 2 == 0 else WHITE
    for ci, val in enumerate(row_data, 1):
        c = wg.cell(row=ri, column=ci, value=val)
        c.font = bf(10); c.fill = fl(bg)
        c.alignment = lft() if ci == 4 else ctr()
        c.border = bd()
    # Impact
    imp = wg.cell(row=ri, column=9, value=f"=G{ri}-H{ri}")
    imp.number_format = "0.0"
    imp.font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
    imp.fill = fl(bg); imp.alignment = ctr(); imp.border = bd()
    # Q7 – Gravité avec SI imbriqué
    grav = wg.cell(row=ri, column=10,
        value=f'=IF(I{ri}<=1,"Faible",IF(AND(I{ri}>=2,I{ri}<=3),"Moyen","Élevé"))')
    grav.font = bf(10, bold=True); grav.fill = fl(bg)
    grav.alignment = ctr(); grav.border = bd()

# Q7 – Mise en forme conditionnelle sur Gravité
rng = f"J{DATA_FIRST}:J{DATA_LAST}"
wg.conditional_formatting.add(rng, FormulaRule(
    formula=['J3="Faible"'],
    fill=fl(GREEN_FILL), font=Font(color=GREEN_DARK, bold=True, name="Calibri")))
wg.conditional_formatting.add(rng, FormulaRule(
    formula=['J3="Moyen"'],
    fill=fl(ORANGE_FILL), font=Font(color=ORANGE_DARK, bold=True, name="Calibri")))
wg.conditional_formatting.add(rng, FormulaRule(
    formula=['J3="Élevé"'],
    fill=fl(RED_FILL), font=Font(color=RED_DARK, bold=True, name="Calibri")))

# Q8 – AutoFilter combiné sur Salle (C) + Type d'incident (D)
wg.auto_filter.ref = f"A2:J{DATA_LAST}"
# Note explicative sur le filtre
note_row = DATA_LAST + 2
wg.merge_cells(f"A{note_row}:J{note_row}")
n = wg.cell(row=note_row, column=1,
    value="ℹ️  Q8 : Filtre combiné activé – utilisez les flèches ▼ sur les colonnes Salle et Type d'Incident pour filtrer simultanément.")
n.font = Font(name="Calibri", size=9, italic=True, color="595959")
n.alignment = lft()

# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 4 – Tableau Croisé Dynamique  (Question 9)
# ══════════════════════════════════════════════════════════════════════════════
# openpyxl ne crée pas de vrais TCD ; on les simule avec COUNTIF / AVERAGEIF
# et on insère un commentaire indiquant comment créer le vrai TCD dans Excel.
wtcd = wb.create_sheet("TCD – Q9")
wtcd.sheet_view.showGridLines = False
wtcd.sheet_properties.tabColor = "7030A0"

wtcd.merge_cells("A1:H1")
c = wtcd["A1"]
c.value     = "📊  Q9 – TABLEAU CROISÉ DYNAMIQUE – Incidents & Impact"
c.font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
c.fill      = fl(DARK_BLUE)
c.alignment = ctr()
wtcd.row_dimensions[1].height = 38

# ── Bloc 1 : Nb d'incidents par Salle ───────────────────────────────────────
section_banner(wtcd, 3, 1, 3, "Nombre d'incidents par Salle", MID_BLUE)
for ci, h in enumerate(["Salle", "Nb Incidents", "% du Total"], 1):
    style_header_cell(wtcd.cell(row=4, column=ci), h, LIGHT_BLUE, 10)
    wtcd.cell(row=4, column=ci).font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)

for j, sal in enumerate(salles, 5):
    bg = ALT_ROW if j % 2 == 0 else WHITE
    c = wtcd.cell(row=j, column=1, value=sal)
    c.font = bf(10); c.fill = fl(bg); c.alignment = ctr(); c.border = bd()
    nb = wtcd.cell(row=j, column=2, value=f'=COUNTIF(Données!C$3:C$22,A{j})')
    nb.number_format = "0"
    nb.font = Font(name="Calibri", size=10, bold=True, color=MID_BLUE)
    nb.fill = fl(bg); nb.alignment = ctr(); nb.border = bd()
    pct = wtcd.cell(row=j, column=3, value=f'=B{j}/COUNTA(Données!C$3:C$22)')
    pct.number_format = "0.0%"
    pct.font = bf(10); pct.fill = fl(bg); pct.alignment = ctr(); pct.border = bd()

# Ligne Total
tot_row = 5 + len(salles)
style_label(wtcd.cell(row=tot_row, column=1), "TOTAL", MID_BLUE)
wtcd.cell(row=tot_row, column=1).font = Font(name="Calibri", size=10, bold=True, color=WHITE)
tot_nb = wtcd.cell(row=tot_row, column=2, value=f"=SUM(B5:B{tot_row-1})")
tot_nb.number_format = "0"
tot_nb.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
tot_nb.fill = fl(MID_BLUE); tot_nb.alignment = ctr(); tot_nb.border = bd()
tot_pct = wtcd.cell(row=tot_row, column=3, value=f"=SUM(C5:C{tot_row-1})")
tot_pct.number_format = "0.0%"
tot_pct.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
tot_pct.fill = fl(MID_BLUE); tot_pct.alignment = ctr(); tot_pct.border = bd()

# ── Bloc 2 : Impact moyen par Filière ───────────────────────────────────────
section_banner(wtcd, 3, 5, 3, "Impact moyen par Filière", MID_BLUE)
for ci, h in enumerate(["Filière", "Impact Moyen", "Nb Étudiants"], 5):
    style_header_cell(wtcd.cell(row=4, column=ci), h, LIGHT_BLUE, 10)
    wtcd.cell(row=4, column=ci).font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)

for j, fil in enumerate(filieres, 5):
    bg = ALT_ROW if j % 2 == 0 else WHITE
    c = wtcd.cell(row=j, column=5, value=fil)
    c.font = bf(10); c.fill = fl(bg); c.alignment = ctr(); c.border = bd()
    avg = wtcd.cell(row=j, column=6,
        value=f'=AVERAGEIF(Données!B$3:B$22,E{j},Données!I$3:I$22)')
    avg.number_format = "0.00"
    avg.font = Font(name="Calibri", size=10, bold=True, color=MID_BLUE)
    avg.fill = fl(bg); avg.alignment = ctr(); avg.border = bd()
    cnt = wtcd.cell(row=j, column=7,
        value=f'=COUNTIF(Données!B$3:B$22,E{j})')
    cnt.number_format = "0"
    cnt.font = bf(10); cnt.fill = fl(bg); cnt.alignment = ctr(); cnt.border = bd()

# Ligne Total filières
tf = 5 + len(filieres)
style_label(wtcd.cell(row=tf, column=5), "TOTAL", MID_BLUE)
wtcd.cell(row=tf, column=5).font = Font(name="Calibri", size=10, bold=True, color=WHITE)
avg_tot = wtcd.cell(row=tf, column=6, value="=AVERAGE(Données!I3:I22)")
avg_tot.number_format = "0.00"
avg_tot.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
avg_tot.fill = fl(MID_BLUE); avg_tot.alignment = ctr(); avg_tot.border = bd()
cnt_tot = wtcd.cell(row=tf, column=7, value="=COUNTA(Données!A3:A22)")
cnt_tot.number_format = "0"
cnt_tot.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
cnt_tot.fill = fl(MID_BLUE); cnt_tot.alignment = ctr(); cnt_tot.border = bd()

for col, w in zip(["A","B","C","D","E","F","G"], [12, 14, 12, 3, 16, 14, 14]):
    wtcd.column_dimensions[col].width = w

# Note explicative TCD
note_r = tf + 2
wtcd.merge_cells(f"A{note_r}:G{note_r}")
nt = wtcd.cell(row=note_r, column=1,
    value="ℹ️  Pour créer un vrai TCD Excel : sélectionnez Données!A2:I22 → Insertion → Tableau croisé dynamique")
nt.font = Font(name="Calibri", size=9, italic=True, color="595959")
nt.alignment = lft()

# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 5 – Graphique  (Question 10)
# ══════════════════════════════════════════════════════════════════════════════
wc = wb.create_sheet("Graphique – Q10")
wc.sheet_view.showGridLines = False
wc.sheet_properties.tabColor = "ED7D31"

wc.merge_cells("A1:F1")
c = wc["A1"]
c.value     = "📊  Q10 – GRAPHIQUE EN COLONNES – Incidents par Type"
c.font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
c.fill      = fl(DARK_BLUE)
c.alignment = ctr()
wc.row_dimensions[1].height = 38

# Table source du graphique
style_header_cell(wc.cell(row=2, column=1), "Type d'Incident")
style_header_cell(wc.cell(row=2, column=2), "Nombre d'Incidents")
wc.column_dimensions["A"].width = 28
wc.column_dimensions["B"].width = 20
wc.row_dimensions[2].height = 26

for j, inc in enumerate(incidents, 3):
    bg = ALT_ROW if j % 2 == 0 else WHITE
    c = wc.cell(row=j, column=1, value=inc)
    c.font = bf(10); c.fill = fl(bg); c.alignment = lft(); c.border = bd()
    v = wc.cell(row=j, column=2, value=f'=COUNTIF(Données!D$3:D$22,A{j})')
    v.number_format = "0"
    v.font = Font(name="Calibri", size=10, bold=True, color=MID_BLUE)
    v.fill = fl(bg); v.alignment = ctr(); v.border = bd()

n = len(incidents)

# Graphique en colonnes
chart = BarChart()
chart.type      = "col"
chart.grouping  = "clustered"
chart.title     = "Nombre d'Incidents par Type de Connexion – SupNum 2025-2026"
chart.y_axis.title = "Nombre d'Incidents"
chart.x_axis.title = "Type d'Incident"
chart.style     = 10
chart.width     = 22
chart.height    = 14

data_ref = Reference(wc, min_col=2, min_row=2, max_row=2 + n)
cats_ref = Reference(wc, min_col=1, min_row=3, max_row=2 + n)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

# Étiquettes de données visibles (Q10 obligatoire)
chart.dataLabels          = DataLabelList()
chart.dataLabels.showVal  = True
chart.dataLabels.showCatName = False

# Couleurs personnalisées par barre
bar_colors = ["2E75B6", "ED7D31", "70AD47", "FF0000", "7030A0"]
for i, series in enumerate(chart.series):
    series.graphicalProperties.solidFill = bar_colors[i % len(bar_colors)]
    series.graphicalProperties.line.solidFill = "FFFFFF"

wc.add_chart(chart, "D3")

# ══════════════════════════════════════════════════════════════════════════════
# Sauvegarde
# ══════════════════════════════════════════════════════════════════════════════
out = "PPP_S3_MATRICULES.xlsx"
wb.save(out)
print(f"✅  Fichier sauvegardé : {out}")
print("Feuilles :", wb.sheetnames)