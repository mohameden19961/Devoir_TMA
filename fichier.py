from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.numbers import FORMAT_NUMBER_00
import random

# ─── Palette ────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
HEADER_BG   = "2E75B6"
ALT_ROW     = "EBF3FB"
WHITE       = "FFFFFF"
ACCENT_GOLD = "F4B942"
GRAY_BORDER = "B8CCE4"
GREEN_OK    = "70AD47"
RED_ERR     = "FF0000"
ORANGE_MED  = "ED7D31"

def make_border(color=GRAY_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ─── Data ────────────────────────────────────────────────────────────────────
filieres   = ["Informatique", "Réseaux", "IA", "Cybersécurité", "Data Science"]
salles     = ["Salle A101", "Salle B202", "Salle C303", "Salle D404", "Salle E505"]
incidents  = ["Coupure Wi-Fi totale", "Connexion instable", "DNS indisponible",
              "Latence élevée", "Déconnexion répétée"]

random.seed(42)

rows = []
for i in range(1, 21):
    filiere  = random.choice(filieres)
    salle    = random.choice(salles)
    incident = random.choice(incidents)
    duree    = random.randint(5, 45)
    tentatives = random.randint(1, 8)
    note_init  = round(random.uniform(12, 20), 1)
    perte      = round(random.uniform(0.5, 5.0), 1)
    note_fin   = round(max(0, note_init - perte), 1)
    rows.append([f"ETU{1000+i}", filiere, salle, incident,
                 duree, tentatives, note_init, note_fin])

# ─── Workbook ────────────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Données
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Données"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A3"

# --- Title block ---
ws.merge_cells("A1:I1")
title_cell = ws["A1"]
title_cell.value = "📊  ANALYSE DES INCIDENTS DE CONNEXION – SupNum 2025-2026"
title_cell.font   = Font(name="Calibri", size=14, bold=True, color=WHITE)
title_cell.fill   = fill(DARK_BLUE)
title_cell.alignment = center()
ws.row_dimensions[1].height = 36

# --- Column headers ---
headers = ["ID Étudiant", "Filière", "Salle", "Type d'Incident",
           "Durée (min)", "Nb Tentatives", "Note Initiale", "Note Finale", "Impact"]
col_widths = [14, 16, 14, 24, 14, 15, 14, 13, 10]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=2, column=col_idx, value=h)
    cell.font      = hdr_font()
    cell.fill      = fill(HEADER_BG)
    cell.alignment = center()
    cell.border    = make_border(MID_BLUE)
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[2].height = 28

# --- Data rows ---
for r_idx, row_data in enumerate(rows, start=3):
    bg = ALT_ROW if r_idx % 2 == 0 else WHITE
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font      = body_font()
        cell.fill      = fill(bg)
        cell.alignment = center() if c_idx != 4 else left()
        cell.border    = make_border()
    # Impact formula  (col I = 9)
    impact_cell = ws.cell(row=r_idx, column=9)
    impact_cell.value      = f"=G{r_idx}-H{r_idx}"
    impact_cell.number_format = "0.0"
    impact_cell.font       = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
    impact_cell.fill       = fill(bg)
    impact_cell.alignment  = center()
    impact_cell.border     = make_border()

# Color-scale on Impact column (I3:I22)
ws.conditional_formatting.add(
    "I3:I22",
    ColorScaleRule(
        start_type="num", start_value=0,   start_color="63BE7B",
        mid_type="num",   mid_value=2.5,   mid_color="FFEB84",
        end_type="num",   end_value=5,     end_color="F8696B"
    )
)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Analyse
# ════════════════════════════════════════════════════════════════════════════
wa = wb.create_sheet("Analyse")
wa.sheet_view.showGridLines = False

def section_title(ws, row, col, text, span=4, bg=DARK_BLUE):
    end_col = get_column_letter(col + span - 1)
    start_col = get_column_letter(col)
    try:
        ws.unmerge_cells(f"{start_col}{row}:{end_col}{row}")
    except Exception:
        pass
    ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", size=12, bold=True, color=WHITE)
    c.fill      = fill(bg)
    c.alignment = center()
    ws.row_dimensions[row].height = 26
    return c

def kpi_row(ws, row, label, formula, col_a=1, fmt="0.0"):
    lc = ws.cell(row=row, column=col_a, value=label)
    lc.font      = body_font(bold=True, color=DARK_BLUE)
    lc.fill      = fill(ALT_ROW)
    lc.alignment = left()
    lc.border    = make_border()

    vc = ws.cell(row=row, column=col_a+1, value=formula)
    vc.number_format = fmt
    vc.font      = Font(name="Calibri", size=10, bold=True, color=MID_BLUE)
    vc.fill      = fill(WHITE)
    vc.alignment = center()
    vc.border    = make_border()

# Title
wa.merge_cells("A1:H1")
t = wa["A1"]
t.value     = "📈  TABLEAU D'ANALYSE – Incidents de Connexion"
t.font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
t.fill      = fill(DARK_BLUE)
t.alignment = center()
wa.row_dimensions[1].height = 36

# --- Section 1: KPI Statistiques ---
section_title(wa, 3, 1, "📌  Statistiques Globales", span=4, bg=MID_BLUE)

kpi_labels = [
    ("Durée moyenne des coupures (min)",    "=AVERAGE(Données!E3:E22)"),
    ("Durée minimale (min)",                "=MIN(Données!E3:E22)"),
    ("Durée maximale (min)",                "=MAX(Données!E3:E22)"),
    ("Impact moyen sur les notes",          "=AVERAGE(Données!I3:I22)"),
    ("Nb étudiants ayant perdu > 2 points", "=COUNTIF(Données!I3:I22,\">2\")"),
]
for i, (lbl, frm) in enumerate(kpi_labels, start=4):
    fmt = "0.00" if "moyen" in lbl or "Impact" in lbl else "0"
    kpi_row(wa, i, lbl, frm, fmt=fmt)

wa.column_dimensions["A"].width = 38
wa.column_dimensions["B"].width = 14

# --- Section 2: Incidents par Type ---
section_title(wa, 10, 1, "📋  Incidents par Type", span=3, bg=MID_BLUE)

inc_types = list(set(incidents))
wa.cell(row=11, column=1, value="Type d'Incident").font = hdr_font(color=DARK_BLUE)
wa.cell(row=11, column=1).fill = fill(LIGHT_BLUE)
wa.cell(row=11, column=1).alignment = center()
wa.cell(row=11, column=2, value="Nombre").font = hdr_font(color=DARK_BLUE)
wa.cell(row=11, column=2).fill = fill(LIGHT_BLUE)
wa.cell(row=11, column=2).alignment = center()

for j, t_name in enumerate(inc_types, start=12):
    wa.cell(row=j, column=1, value=t_name).border = make_border()
    wa.cell(row=j, column=1).font = body_font()
    wa.cell(row=j, column=1).fill = fill(ALT_ROW if j%2==0 else WHITE)
    cnt_cell = wa.cell(row=j, column=2,
                       value=f'=COUNTIF(Données!D3:D22,A{j})')
    cnt_cell.border    = make_border()
    cnt_cell.alignment = center()
    cnt_cell.font      = Font(name="Calibri", size=10, bold=True, color=MID_BLUE)
    cnt_cell.fill      = fill(ALT_ROW if j%2==0 else WHITE)

# --- Section 3: Incidents par Salle ---
section_title(wa, 10, 4, "🏫  Incidents par Salle", span=3, bg=MID_BLUE)
wa.cell(row=11, column=4, value="Salle").font = hdr_font(color=DARK_BLUE)
wa.cell(row=11, column=4).fill = fill(LIGHT_BLUE)
wa.cell(row=11, column=4).alignment = center()
wa.cell(row=11, column=5, value="Nb Incidents").font = hdr_font(color=DARK_BLUE)
wa.cell(row=11, column=5).fill = fill(LIGHT_BLUE)
wa.cell(row=11, column=5).alignment = center()

for j, salle in enumerate(salles, start=12):
    wa.cell(row=j, column=4, value=salle).border = make_border()
    wa.cell(row=j, column=4).font = body_font()
    wa.cell(row=j, column=4).fill = fill(ALT_ROW if j%2==0 else WHITE)
    cnt = wa.cell(row=j, column=5,
                  value=f'=COUNTIF(Données!C3:C22,D{j})')
    cnt.border    = make_border()
    cnt.alignment = center()
    cnt.font      = Font(name="Calibri", size=10, bold=True, color=MID_BLUE)
    cnt.fill      = fill(ALT_ROW if j%2==0 else WHITE)

wa.column_dimensions["D"].width = 18
wa.column_dimensions["E"].width = 14

# Salle la plus touchée
wa.merge_cells("A19:B19")
lbl = wa["A19"]
lbl.value     = "🏆 Salle la plus touchée :"
lbl.font      = Font(name="Calibri", size=10, bold=True, color=WHITE)
lbl.fill      = fill(ACCENT_GOLD)
lbl.alignment = left()
lbl.border    = make_border()

wa.merge_cells("C19:E19")
val = wa["C19"]
val.value     = f'=INDEX(D12:D{11+len(salles)},MATCH(MAX(E12:E{11+len(salles)}),E12:E{11+len(salles)},0))'
val.font      = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
val.fill      = fill(LIGHT_BLUE)
val.alignment = center()
val.border    = make_border()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Gravité & Filtre
# ════════════════════════════════════════════════════════════════════════════
wg = wb.create_sheet("Gravité")
wg.sheet_view.showGridLines = False
wg.freeze_panes = "A3"

wg.merge_cells("A1:J1")
t = wg["A1"]
t.value     = "🚦  ANALYSE DE LA GRAVITÉ DES INCIDENTS"
t.font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
t.fill      = fill(DARK_BLUE)
t.alignment = center()
wg.row_dimensions[1].height = 36

headers_g = ["ID Étudiant", "Filière", "Salle", "Type d'Incident",
             "Durée (min)", "Nb Tentatives", "Note Initiale", "Note Finale", "Impact", "Gravité"]
col_widths_g = [14, 16, 14, 24, 13, 14, 14, 13, 10, 12]

for ci, (h, w) in enumerate(zip(headers_g, col_widths_g), start=1):
    c = wg.cell(row=2, column=ci, value=h)
    c.font      = hdr_font()
    c.fill      = fill(HEADER_BG)
    c.alignment = center()
    c.border    = make_border(MID_BLUE)
    wg.column_dimensions[get_column_letter(ci)].width = w
wg.row_dimensions[2].height = 28

# Data with Gravité formula
for r, row_data in enumerate(rows, start=3):
    bg = ALT_ROW if r % 2 == 0 else WHITE
    for ci, val in enumerate(row_data, start=1):
        c = wg.cell(row=r, column=ci, value=val)
        c.font = body_font()
        c.fill = fill(bg)
        c.alignment = center() if ci != 4 else left()
        c.border = make_border()

    # Impact (col 9)
    imp = wg.cell(row=r, column=9, value=f"=G{r}-H{r}")
    imp.number_format = "0.0"
    imp.font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
    imp.fill = fill(bg)
    imp.alignment = center()
    imp.border = make_border()

    # Gravité (col 10) – SI formula
    grav = wg.cell(row=r, column=10,
                   value=f'=IF(I{r}<=1,"Faible",IF(AND(I{r}>=2,I{r}<=3),"Moyen","Élevé"))')
    grav.font = body_font(bold=True)
    grav.fill = fill(bg)
    grav.alignment = center()
    grav.border = make_border()

# Conditional formatting on Gravité column
wg.conditional_formatting.add(
    f"J3:J{2+len(rows)}",
    FormulaRule(formula=['J3="Faible"'], fill=PatternFill("solid", fgColor="C6EFCE"),
                font=Font(color="276221", bold=True, name="Calibri"))
)
wg.conditional_formatting.add(
    f"J3:J{2+len(rows)}",
    FormulaRule(formula=['J3="Moyen"'], fill=PatternFill("solid", fgColor="FFEB9C"),
                font=Font(color="9C6500", bold=True, name="Calibri"))
)
wg.conditional_formatting.add(
    f"J3:J{2+len(rows)}",
    FormulaRule(formula=['J3="Élevé"'], fill=PatternFill("solid", fgColor="FFC7CE"),
                font=Font(color="9C0006", bold=True, name="Calibri"))
)

# Enable AutoFilter
wg.auto_filter.ref = f"A2:J{2+len(rows)}"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 – Graphique (chart data)
# ════════════════════════════════════════════════════════════════════════════
wc = wb.create_sheet("Graphique")
wc.sheet_view.showGridLines = False

wc.merge_cells("A1:F1")
t = wc["A1"]
t.value     = "📊  GRAPHIQUE – Incidents par Type"
t.font      = Font(name="Calibri", size=14, bold=True, color=WHITE)
t.fill      = fill(DARK_BLUE)
t.alignment = center()
wc.row_dimensions[1].height = 36

# Header
for ci, hdr in enumerate(["Type d'Incident", "Nombre d'Incidents"], start=1):
    c = wc.cell(row=2, column=ci, value=hdr)
    c.font = hdr_font()
    c.fill = fill(HEADER_BG)
    c.alignment = center()
    c.border = make_border()
wc.column_dimensions["A"].width = 28
wc.column_dimensions["B"].width = 18

# Data linked from Analyse sheet
for j, inc in enumerate(inc_types, start=3):
    ca = wc.cell(row=j, column=1, value=inc)
    ca.font = body_font()
    ca.fill = fill(ALT_ROW if j%2==0 else WHITE)
    ca.border = make_border()
    ca.alignment = left()

    cb = wc.cell(row=j, column=2,
                 value=f'=COUNTIF(Données!D3:D22,A{j})')
    cb.font = Font(name="Calibri", size=10, bold=True, color=MID_BLUE)
    cb.fill = fill(ALT_ROW if j%2==0 else WHITE)
    cb.border = make_border()
    cb.alignment = center()

n_types = len(inc_types)

# Bar chart
chart = BarChart()
chart.type    = "col"
chart.grouping = "clustered"
chart.title   = "Nombre d'Incidents par Type de Connexion"
chart.y_axis.title = "Nombre d'Incidents"
chart.x_axis.title = "Type d'Incident"
chart.style   = 10
chart.width   = 20
chart.height  = 14

data_ref   = Reference(wc, min_col=2, min_row=2, max_row=2+n_types)
cats_ref   = Reference(wc, min_col=1, min_row=3, max_row=2+n_types)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True

# Color the bars
from openpyxl.chart.data_source import NumDataSource, NumRef
colors_list = ["2E75B6", "ED7D31", "A9D18E", "FF0000", "7030A0"]
for i, series in enumerate(chart.series):
    series.graphicalProperties.solidFill = colors_list[i % len(colors_list)]
    series.graphicalProperties.line.solidFill = "FFFFFF"

wc.add_chart(chart, "D3")

# ════════════════════════════════════════════════════════════════════════════
# Tab colors
# ════════════════════════════════════════════════════════════════════════════
ws.sheet_properties.tabColor  = "1F3864"
wa.sheet_properties.tabColor  = "2E75B6"
wg.sheet_properties.tabColor  = "70AD47"
wc.sheet_properties.tabColor  = "ED7D31"

# ────────────────────────────────────────────────────────────────────────────
out = "/mnt/user-data/outputs/PPP_S3_MATRICULES.xlsx"
wb.save(out)
print(f"✅  Fichier sauvegardé : {out}")